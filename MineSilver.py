from bs4 import BeautifulSoup
import requests as req
import xml.etree.ElementTree as ET
import openpyxl as pyxl
import os
import string
import re
from transliterate import translit
import dateparser as DP
import unicodedata


fname = 'metadata.xlsx'                                    # file name

corepath = f'{os.getcwd()}\\SilverAgePoets_data'           # путь для папки со всеми данными
metapath = f'{corepath}\\{fname}'                          # путь для Excel-файла

columnsnames = ['id', 'author', 'title', 'year',
                'trend', 'genre', 'birth', 'wordcount']    # имена колонок в Excel-файле
poets = [
    {'author': 'Блок', 'address': 'blok', 'birth': '1880', 'trend': 'symbolism', 'genre': 'undef'},
    {'author': 'Бунин', 'address': 'bunin', 'birth': '1870', 'trend': 'symbolism', 'genre': 'undef'},
    {'author': 'Брюсов', 'address': 'briusov', 'birth': '1873', 'trend': 'symbolism', 'genre': 'undef'},
    {'author': 'Мережковский', 'address': 'merezhkovskiy', 'birth': '1865', 'trend': 'symbolism', 'genre': 'undef'},
    {'author': 'Гиппиус', 'address': 'gippiusz', 'birth': '1869', 'trend': 'symbolism', 'genre': 'undef'},
    {'author': 'Сологуб', 'address': 'sologub', 'birth': '1863', 'trend': 'symbolism', 'genre': 'undef'},
    {'author': 'Г. Иванов', 'address': 'ivanovg', 'birth': '1894', 'trend': 'acmeism', 'genre': 'undef'},
    {'author': 'Мандельштам', 'address': 'mandelshtam', 'birth': '1891', 'trend': 'acmeism', 'genre': 'undef'},
    {'author': 'Адамович', 'address': 'adamovich', 'birth': '1892', 'trend': 'acmeism', 'genre': 'undef'},
    {'author': 'Маяковский', 'address': 'mayakovskiy', 'birth': '1893', 'trend': 'futurism', 'genre': 'undef'},
    {'author': 'Хлебников', 'address': 'hlebnikov', 'birth': '1885', 'trend': 'futurism', 'genre': 'undef'},
    {'author': 'Есенин', 'address': 'esenin', 'birth': '1895', 'trend': 'imaginism', 'genre': 'undef'},
    {'author': 'Шершеневич', 'address': 'shershenevich', 'birth': '1893', 'trend': 'imaginism', 'genre': 'undef'},
    {'author': 'Цветаева', 'address': 'cvetaeva', 'birth': '1892', 'trend': 'undef', 'genre': 'undef'},
    {'author': 'Ходасевич', 'address': 'hodasevich', 'birth': '1886', 'trend': 'undef', 'genre': 'undef'},
    {'author': 'Хармс', 'address': 'harms', 'birth': '1905', 'trend': 'undef', 'genre': 'undef'},
    {'author': 'Бедный', 'address': 'bednyy', 'birth': '1883', 'trend': 'socrealism', 'genre': 'undef'},
    {'author': 'Вертинский', 'address': 'vertinskij', 'birth': '1889', 'trend': 'undef', 'genre': 'undef'},
    {'author': 'Заболоцкий', 'address': 'zabolotskiy', 'birth': '1903', 'trend': 'undef', 'genre': 'undef'},
    {'author': 'Кузмин', 'address': 'kuzmin', 'birth': '1872', 'trend': 'undef', 'genre': 'undef'},
]


def parse4poems(p):
    tlink = 'https://slova.org.ru/{}/'
    for poet in p:
        print(f'* * *  {poet["author"]}  * * *')

        author_addr = poet['address']
        author = poet['author']
        kauthor = translit(author, reversed=True)
        birth = poet['birth']
        trend = poet['trend']
        genre = poet['genre']

        # authorpath = f'{corepath}\\{author}'
        # filepath = f'{authorpath}\\files'

        # os.mkdir(authorpath)
        # os.chdir(authorpath)
        # os.mkdir(filepath)

        ftl = tlink.format(f'{author_addr}/index')
        resp = req.get(ftl)
        soup = BeautifulSoup(resp.text, 'lxml')
        poems = soup.find('div', id='stihi_list')

        links = [tag['href'] for tag in poems.find_all('a')]

        copies = dict()
        wnumber = 0

        print(
            f'Начинаем обработку.\n'
            f'Всего ссылок: {len(links)}\n'
        )
        try:
            for link in links:
                l = tlink.format(link)
                poem_resp = req.get(l)
                poem_page = BeautifulSoup(poem_resp.text, 'lxml')

                text_tag = poem_page.find('pre')
                ptxt = text_tag.text
                try:
                    poem_lines = [line.split('\r')[0]
                                  for line in ptxt.split('\n')
                                  if line.split('\r')[0] != '']
                except:
                    poem_lines = [line
                                  for line in ptxt.split('\n')
                                  if line != '']

                year_tag = text_tag.find('i')
                try:
                    year = str(DP.parse(year_tag.text).year) if DP.parse(year_tag.text) is not None else 'undef'
                except:
                    year = 'undef'

                unpnctor = re.compile('[%s]' % re.escape(string.punctuation + '—'))

                title = poem_page.find('h3').text
                title = unpnctor.sub('', title)

                copies[title] = 1 if title not in copies.keys() else copies[title] + 1
                if copies[title] > 1:
                    teaser = ''.join([ch for ch in poem_lines[0] if unicodedata.category(ch) != 'Cc'])
                    teaser = unpnctor.sub('', teaser)
                    title = f'{title}_{teaser}'
                else:
                    title = title

                text_unpncted = unpnctor.sub('', ptxt)
                wordcount = len(text_unpncted.split())
                wnumber += wordcount

                try:
                    ktitle = translit(title, reversed=True)
                except:
                    ktitle = title

                idpoem = f'{year}_{kauthor}_{ktitle}'
                data = [idpoem, author, title, year,
                        trend, genre, birth, wordcount]

                root = ET.Element('poem')
                tree = ET.ElementTree(root)
                ET.SubElement(
                    root,
                    'meta',
                    author=author, year=year,
                    trend=trend, genre=genre,
                    nwords=str(wordcount), source=l
                )
                b = ET.SubElement(root, 'body')
                ET.SubElement(b, 'title').text = title
                t = ET.SubElement(b, 'text')

                linenum = '1'
                for line in poem_lines:
                    ET.SubElement(t, 'lb', n=linenum).text = line
                    linenum = str(int(linenum) + 1)

                # os.chdir(filepath)
                # tree.write(f'{filepath}\\{ktitle}.xml', 'utf-8')
                # os.chdir(authorpath)
                tree.write(f'{corepath}\\{idpoem}.xml', 'utf-8')
                print(
                    f'~~~\n'
                    f'Название: {title}\n'
                    f'Количество слов: {wordcount}\n'
                    f'Осталось ссылок: {len(links) - links.index(link) - 1}'
                )
                wbsheet.append(data)
            wb.save(metapath)
            print(f'\n"{author}" - общее количество слов: {wnumber}\n')
        except:
            wb.save(metapath)
            print(f'\n"{author}" - последнее стиховторение: {title}\n')
            print('- Choto went wrong!\n'
                  '- Cho?\n'
                  '- Nie znayou...')
            raise


if 'SilverAgePoets_data' not in os.listdir() and os.getcwd() != corepath:
    os.mkdir(corepath)
    os.chdir(corepath)

    wb = pyxl.Workbook()
    wbsheet = wb.active
    wbsheet.append(columnsnames)
    wb.save(fname)

    parse4poems(poets)
else:
    os.chdir(corepath)

    wb = pyxl.load_workbook(fname)
    sheets = wb.sheetnames
    wbsheet = wb[sheets[0]]

    parse4poems(poets)
